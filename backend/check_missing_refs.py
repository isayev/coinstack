"""Check coins without references."""
import sys
sys.path.insert(0, '.')

import openpyxl
from app.database import SessionLocal
from app.models import Coin, CoinReference
from app.services.excel_import import normalize_text

wb = openpyxl.load_workbook(r'c:\vibecode\coinstack\original-data\collection-v1.xlsx')
ws = wb.active

# Get headers
headers = {}
for col_idx, cell in enumerate(ws[1], start=1):
    if cell.value:
        headers[normalize_text(cell.value)] = col_idx

db = SessionLocal()

# Find coins without references
coins_without_refs = db.query(Coin).outerjoin(CoinReference).filter(CoinReference.id == None).all()

print(f"Coins without references in DB: {len(coins_without_refs)}")
print("=" * 80)

# Check Excel for these coins
for coin in coins_without_refs[:20]:
    # Find in Excel
    for row in range(2, 120):
        ruler = normalize_text(ws.cell(row, headers.get("Ruler Issuer")).value)
        denom = normalize_text(ws.cell(row, headers.get("Coin type")).value)
        ref = normalize_text(ws.cell(row, headers.get("Reference")).value)
        
        if ruler == coin.issuing_authority and denom == coin.denomination:
            print(f"\n{coin.issuing_authority} - {coin.denomination}")
            print(f"  Excel Reference: '{ref if ref else 'EMPTY'}'")
            break

db.close()
