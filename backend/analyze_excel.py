import openpyxl

wb = openpyxl.load_workbook(r'c:\vibecode\coinstack\original-data\collection-v1.xlsx')
ws = wb.active

# Get headers
print("=" * 80)
print("EXCEL HEADERS (Column mapping analysis)")
print("=" * 80)
headers = []
for col_idx, cell in enumerate(ws[1], start=1):
    header = str(cell.value).strip() if cell.value else f"<empty col {col_idx}>"
    headers.append(header)
    print(f"  Col {col_idx}: '{header}'")

print("\n" + "=" * 80)
print("SAMPLE DATA (First 3 data rows)")
print("=" * 80)

for row_idx in range(2, 5):
    print(f"\n--- Row {row_idx} ---")
    for col_idx, cell in enumerate(ws[row_idx], start=1):
        if col_idx <= len(headers):
            header = headers[col_idx - 1]
            value = cell.value
            if value is not None and str(value).strip():
                print(f"  {header}: '{value}'")

# Check all unique values for key columns
print("\n" + "=" * 80)
print("UNIQUE VALUES IN KEY COLUMNS")
print("=" * 80)

# Find column indices
header_map = {h: i for i, h in enumerate(headers)}

def get_unique_values(col_name):
    if col_name not in header_map:
        return []
    col_idx = header_map[col_name] + 1
    values = set()
    for row in ws.iter_rows(min_row=2, max_col=col_idx, min_col=col_idx, values_only=True):
        if row[0]:
            values.add(str(row[0]).strip())
    return sorted(values)

key_columns = ["Category", "Composition", "Minted", "Ruled", "Condition", "NGC Grade", "Reference"]
for col in key_columns:
    values = get_unique_values(col)
    print(f"\n{col} ({len(values)} unique):")
    for v in values[:20]:  # Show first 20
        print(f"  - '{v}'")
    if len(values) > 20:
        print(f"  ... and {len(values) - 20} more")
