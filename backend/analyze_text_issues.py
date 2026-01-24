"""Analyze text issues in the Excel file."""
import openpyxl
import unicodedata
import re

wb = openpyxl.load_workbook(r'c:\vibecode\coinstack\original-data\collection-v1.xlsx')
ws = wb.active

# Get headers
headers = []
for cell in ws[1]:
    headers.append(str(cell.value).strip() if cell.value else "")

print("=" * 80)
print("ANALYZING TEXT ISSUES IN EXCEL FILE")
print("=" * 80)

# Collect all text values and analyze
issues = {
    "non_ascii": [],
    "multiple_spaces": [],
    "special_dashes": [],
    "leading_trailing_spaces": [],
    "special_quotes": [],
    "control_chars": [],
    "nbsp": [],
}

def analyze_text(value, row, col_name):
    if not value:
        return
    text = str(value)
    
    # Check for non-ASCII
    non_ascii = [c for c in text if ord(c) > 127]
    if non_ascii:
        unique_chars = set(non_ascii)
        issues["non_ascii"].append({
            "row": row, "col": col_name,
            "chars": [(c, hex(ord(c)), unicodedata.name(c, "UNKNOWN")) for c in unique_chars],
            "sample": text[:60]
        })
    
    # Check for multiple spaces
    if "  " in text:
        issues["multiple_spaces"].append({
            "row": row, "col": col_name,
            "sample": text[:60]
        })
    
    # Check for special dashes (en-dash, em-dash, minus sign)
    special_dashes = re.findall(r'[–—−]', text)
    if special_dashes:
        issues["special_dashes"].append({
            "row": row, "col": col_name,
            "dashes": [(d, hex(ord(d)), unicodedata.name(d, "UNKNOWN")) for d in set(special_dashes)],
            "sample": text[:60]
        })
    
    # Check for leading/trailing spaces
    if text != text.strip():
        issues["leading_trailing_spaces"].append({
            "row": row, "col": col_name,
            "sample": repr(text[:40])
        })
    
    # Check for special quotes
    special_quotes = re.findall(r'[""''„‟‹›«»]', text)
    if special_quotes:
        issues["special_quotes"].append({
            "row": row, "col": col_name,
            "quotes": list(set(special_quotes)),
            "sample": text[:60]
        })
    
    # Check for control characters
    control_chars = [c for c in text if ord(c) < 32 and c not in '\n\r\t']
    if control_chars:
        issues["control_chars"].append({
            "row": row, "col": col_name,
            "chars": [(c, hex(ord(c))) for c in set(control_chars)],
            "sample": repr(text[:60])
        })
    
    # Check for non-breaking spaces
    if '\u00a0' in text or '\u202f' in text:
        issues["nbsp"].append({
            "row": row, "col": col_name,
            "sample": text[:60]
        })

# Analyze all cells
for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
    for col_idx, value in enumerate(row):
        if col_idx < len(headers) and value:
            analyze_text(value, row_idx, headers[col_idx])

# Report findings
for issue_type, items in issues.items():
    print(f"\n{issue_type.upper().replace('_', ' ')} ({len(items)} occurrences):")
    print("-" * 60)
    if not items:
        print("  None found")
    else:
        for item in items[:10]:  # Show first 10
            print(f"  Row {item['row']}, Col '{item['col']}':")
            if 'chars' in item:
                print(f"    Chars: {item['chars']}")
            if 'dashes' in item:
                print(f"    Dashes: {item['dashes']}")
            if 'quotes' in item:
                print(f"    Quotes: {item['quotes']}")
            print(f"    Sample: {item['sample']}")
        if len(items) > 10:
            print(f"  ... and {len(items) - 10} more")

print("\n" + "=" * 80)
print("UNIQUE NON-ASCII CHARACTERS FOUND")
print("=" * 80)

all_non_ascii = set()
for row in ws.iter_rows(min_row=2, values_only=True):
    for value in row:
        if value:
            text = str(value)
            for c in text:
                if ord(c) > 127:
                    all_non_ascii.add(c)

for c in sorted(all_non_ascii, key=ord):
    print(f"  '{c}' (U+{ord(c):04X}) - {unicodedata.name(c, 'UNKNOWN')}")
