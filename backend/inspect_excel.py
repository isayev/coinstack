"""Inspect Excel file structure and test import."""
import sys
from pathlib import Path
import openpyxl

def inspect_excel_file(file_path: Path):
    """Inspect Excel file structure."""
    print(f"Inspecting: {file_path}")
    print("=" * 60)
    
    if not file_path.exists():
        print(f"Error: File not found: {file_path}")
        return
    
    workbook = openpyxl.load_workbook(file_path, data_only=True)
    sheet = workbook.active
    
    # Read headers
    headers = []
    for cell in sheet[1]:
        headers.append(str(cell.value).strip() if cell.value else "")
    
    print(f"\nFound {len(headers)} columns:")
    print("-" * 60)
    for i, header in enumerate(headers, 1):
        print(f"{i:2d}. {header}")
    
    # Read first few data rows
    print(f"\n\nFirst 3 data rows:")
    print("=" * 60)
    
    for row_idx, row in enumerate(sheet.iter_rows(min_row=2, max_row=4, values_only=True), start=2):
        print(f"\nRow {row_idx}:")
        for i, value in enumerate(row):
            if i < len(headers) and headers[i]:
                print(f"  {headers[i]}: {value}")
    
    # Count total rows
    total_rows = sum(1 for _ in sheet.iter_rows(min_row=2)) if sheet.max_row > 1 else 0
    print(f"\n\nTotal data rows: {total_rows}")
    
    workbook.close()

def test_column_mapping():
    """Test the current column mapping."""
    # Import without database dependency
    import sys
    sys.path.insert(0, str(Path(__file__).parent))
    
    from app.services.excel_import import COLUMN_MAPPING, read_excel_file
    
    file_path = Path("../collection-v1.xlsx")
    
    if not file_path.exists():
        print(f"Error: File not found: {file_path}")
        return
    
    print("\n\nTesting Column Mapping:")
    print("=" * 60)
    
    rows = read_excel_file(file_path)
    if not rows:
        print("No rows found!")
        return
    
    # Get all column names from Excel
    excel_columns = set()
    for row in rows[:10]:  # Check first 10 rows
        excel_columns.update(row.keys())
    
    print(f"\nExcel columns found: {len(excel_columns)}")
    print("-" * 60)
    for col in sorted(excel_columns):
        mapped = COLUMN_MAPPING.get(col, "NOT MAPPED")
        status = "✓" if mapped != "NOT MAPPED" else "✗"
        print(f"{status} {col:30s} -> {mapped}")
    
    # Show unmapped columns
    unmapped = [col for col in excel_columns if col not in COLUMN_MAPPING]
    if unmapped:
        print(f"\n⚠️  Unmapped columns ({len(unmapped)}):")
        for col in sorted(unmapped):
            print(f"   - {col}")
    
    # Show mapping coverage
    mapped_count = len([col for col in excel_columns if col in COLUMN_MAPPING])
    print(f"\nMapping coverage: {mapped_count}/{len(excel_columns)} columns")

if __name__ == "__main__":
    excel_path = Path("../collection-v1.xlsx")
    
    if len(sys.argv) > 1:
        excel_path = Path(sys.argv[1])
    
    inspect_excel_file(excel_path)
    test_column_mapping()
