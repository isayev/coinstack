"""Analyze the 'other' references to identify patterns that could be added."""
import sys
sys.path.insert(0, '.')

from collections import Counter
import re
from app.services.excel_import import parse_references, normalize_text
import openpyxl

wb = openpyxl.load_workbook(r'c:\vibecode\coinstack\original-data\collection-v1.xlsx')
ws = wb.active

# Get Reference column
headers = {}
for col_idx, cell in enumerate(ws[1], start=1):
    if cell.value:
        headers[normalize_text(cell.value)] = col_idx

ref_col = headers.get("Reference")

# Collect all 'other' references
others = []
for row in range(2, 120):
    ref = normalize_text(ws.cell(row, ref_col).value)
    if not ref:
        continue
    
    parsed = parse_references(ref)
    for p in parsed:
        if p["system"] == "other":
            others.append(p["number"])

print("=" * 80)
print(f"ANALYZING {len(others)} 'OTHER' REFERENCES")
print("=" * 80)

# Group by pattern
patterns = {
    "RIC with hyphen": [],      # RIC II.3-1514
    "BMCRE with hyphen": [],    # BMCRE-842
    "RSC with hyphen": [],      # RSC-804
    "RIC with parens": [],      # RIC III (Marcus Aurelius) 430
    "RIC volume only": [],      # RIC III, RIC IVi
    "MIR refs": [],             # MIR 18
    "RPC decimal vol": [],      # RPC V.3
    "Page references": [],       # page 258
    "Unknown catalogs": [],
    "Other": [],
}

for ref in others:
    if re.match(r"RIC\s+[IVX]+\.?\d*-\d+", ref, re.IGNORECASE):
        patterns["RIC with hyphen"].append(ref)
    elif re.match(r"BMCRE?-\d+", ref, re.IGNORECASE):
        patterns["BMCRE with hyphen"].append(ref)
    elif re.match(r"RSC-\d+", ref, re.IGNORECASE):
        patterns["RSC with hyphen"].append(ref)
    elif re.match(r"RIC\s+[IVX]+\s*\([^)]+\)\s*\d+", ref, re.IGNORECASE):
        patterns["RIC with parens"].append(ref)
    elif re.match(r"RIC\s+[IVX]+[a-z]*$", ref, re.IGNORECASE):
        patterns["RIC volume only"].append(ref)
    elif re.match(r"MIR\s+\d+", ref, re.IGNORECASE):
        patterns["MIR refs"].append(ref)
    elif re.match(r"RPC\s+V\.\d+$", ref, re.IGNORECASE):
        patterns["RPC decimal vol"].append(ref)
    elif "page" in ref.lower():
        patterns["Page references"].append(ref)
    elif re.match(r"[A-Za-z]+\s+\d+", ref):
        patterns["Unknown catalogs"].append(ref)
    else:
        patterns["Other"].append(ref)

for pattern_name, refs in patterns.items():
    if refs:
        print(f"\n{pattern_name} ({len(refs)}):")
        for ref in sorted(set(refs))[:10]:
            print(f"  - '{ref}'")
        if len(refs) > 10:
            print(f"  ... and {len(set(refs)) - 10} more unique")

# Frequency of unknown catalogs
print("\n" + "=" * 80)
print("UNKNOWN CATALOG PREFIXES")
print("=" * 80)
catalog_prefixes = Counter()
for ref in patterns["Unknown catalogs"]:
    match = re.match(r"([A-Za-z]+)", ref)
    if match:
        catalog_prefixes[match.group(1)] += 1

for prefix, count in catalog_prefixes.most_common(20):
    print(f"  {prefix}: {count}")
