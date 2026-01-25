"""Debug Decius coin reference issue."""
import sys
sys.path.insert(0, '.')

import openpyxl
from app.database import SessionLocal
from app.models import Coin, CoinReference, ReferenceType
from app.services.excel_import import normalize_text, parse_references

wb = openpyxl.load_workbook(r'c:\vibecode\coinstack\original-data\collection-v1.xlsx')
ws = wb.active

# Get headers
headers = {}
for col_idx, cell in enumerate(ws[1], start=1):
    if cell.value:
        headers[normalize_text(cell.value)] = col_idx

db = SessionLocal()

print("Looking for Decius coins:")
print("=" * 80)

# Find Decius in Excel
for row in range(2, 120):
    ruler = normalize_text(ws.cell(row, headers.get("Ruler Issuer")).value)
    if ruler and "Decius" in ruler:
        denom = normalize_text(ws.cell(row, headers.get("Coin type")).value)
        ref = normalize_text(ws.cell(row, headers.get("Reference")).value)
        print(f"\nRow {row}: {ruler} - {denom}")
        print(f"  Reference: '{ref}'")
        print(f"  Parsed: {parse_references(ref) if ref else 'N/A'}")

print("\n" + "=" * 80)
print("Decius coins in DB:")
print("=" * 80)

decius_coins = db.query(Coin).filter(Coin.issuing_authority.like('%Decius%')).all()
for coin in decius_coins:
    refs = db.query(CoinReference).filter(CoinReference.coin_id == coin.id).all()
    print(f"\n{coin.id}: {coin.issuing_authority} - {coin.denomination}")
    print(f"  References: {len(refs)}")
    for ref in refs:
        ref_type = db.query(ReferenceType).filter(ReferenceType.id == ref.reference_type_id).first()
        if ref_type:
            print(f"    - {ref_type.system}: {ref_type.local_ref}")

db.close()
