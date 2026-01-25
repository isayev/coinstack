"""Check all unique Reference values from Excel."""
import openpyxl
import sys
sys.path.insert(0, '.')

from app.services.excel_import import parse_references

wb = openpyxl.load_workbook(r'c:\vibecode\coinstack\original-data\collection-v1.xlsx')
ws = wb.active

# Find Reference column
header_map = {}
for col_idx, cell in enumerate(ws[1], start=1):
    if cell.value:
        header_map[str(cell.value).strip()] = col_idx

ref_col = header_map.get("Reference")
if not ref_col:
    print("Reference column not found!")
    sys.exit(1)

# Get all unique references
refs = set()
for row in ws.iter_rows(min_row=2, min_col=ref_col, max_col=ref_col, values_only=True):
    if row[0]:
        refs.add(str(row[0]).strip())

print(f"Found {len(refs)} unique reference strings\n")
print("=" * 80)

# Parse each and check for issues
issues = []
for ref_str in sorted(refs):
    parsed = parse_references(ref_str)
    
    # Check for "other" entries that might be missed patterns
    others = [p for p in parsed if p["system"] == "other"]
    
    if others:
        issues.append((ref_str, parsed, others))

print(f"References with 'other' entries ({len(issues)}):\n")
for ref_str, parsed, others in issues:
    print(f"Input: '{ref_str}'")
    print(f"Parsed: {parsed}")
    print(f"Others: {[o['number'] for o in others]}")
    print()
