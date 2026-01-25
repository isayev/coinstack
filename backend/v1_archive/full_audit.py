"""Full audit: Compare Excel data vs imported database data field by field."""
import sys
sys.path.insert(0, '.')

import openpyxl
from app.database import SessionLocal
from app.models import Coin, CoinReference, ReferenceType
from app.services.excel_import import normalize_text, parse_reign_years, parse_references, normalize_category, normalize_metal

wb = openpyxl.load_workbook(r'c:\vibecode\coinstack\original-data\collection-v1.xlsx')
ws = wb.active

# Get headers
headers = {}
for col_idx, cell in enumerate(ws[1], start=1):
    if cell.value:
        headers[normalize_text(cell.value)] = col_idx

db = SessionLocal()

print("=" * 100)
print("FULL IMPORT AUDIT - FIELD BY FIELD ANALYSIS")
print("=" * 100)

# Track issues
issues = []

def get_cell(row, col_name):
    if col_name not in headers:
        return None
    return normalize_text(ws.cell(row, headers[col_name]).value)

def add_issue(row, field, problem, excel_val, db_val=None):
    issues.append({
        "row": row,
        "field": field,
        "problem": problem,
        "excel": excel_val,
        "db": db_val
    })

# Get all coins from DB
coins = db.query(Coin).all()
coin_map = {}  # Map by some identifier

print(f"\nDatabase has {len(coins)} coins")
print(f"Excel has ~110 data rows\n")

# ============================================================================
# SECTION 1: DATE PARSING AUDIT
# ============================================================================
print("=" * 100)
print("1. DATE PARSING AUDIT (Minted column)")
print("=" * 100)

minted_issues = []
for row in range(2, 120):
    minted = get_cell(row, "Minted")
    if not minted or minted == "None":
        continue
    
    start, end = parse_reign_years(minted)
    
    # Check for potential issues
    if start is None and end is None:
        minted_issues.append((row, minted, "FAILED TO PARSE", None, None))
    elif start is not None and end is not None:
        # Sanity checks
        if start > end:
            minted_issues.append((row, minted, "START > END", start, end))
        if start > 500 or (end is not None and end > 500):
            minted_issues.append((row, minted, "YEAR > 500 (suspicious)", start, end))
        if start < -200:
            minted_issues.append((row, minted, "YEAR < -200 BC (suspicious)", start, end))
    
    # Show all parsed dates
    print(f"  Row {row:3d}: '{minted}' -> ({start}, {end})")

if minted_issues:
    print(f"\n  ISSUES FOUND ({len(minted_issues)}):")
    for row, val, problem, start, end in minted_issues:
        print(f"    Row {row}: '{val}' - {problem} -> ({start}, {end})")
else:
    print("\n  No date parsing issues found!")

# ============================================================================
# SECTION 2: RULED (REIGN) PARSING AUDIT  
# ============================================================================
print("\n" + "=" * 100)
print("2. RULED (REIGN) PARSING AUDIT")
print("=" * 100)

ruled_issues = []
for row in range(2, 120):
    ruled = get_cell(row, "Ruled")
    if not ruled or ruled == "None":
        continue
    
    start, end = parse_reign_years(ruled)
    
    if start is None and end is None:
        ruled_issues.append((row, ruled, "FAILED TO PARSE"))
    
    print(f"  Row {row:3d}: '{ruled}' -> ({start}, {end})")

if ruled_issues:
    print(f"\n  ISSUES FOUND ({len(ruled_issues)}):")
    for row, val, problem in ruled_issues:
        print(f"    Row {row}: '{val}' - {problem}")
else:
    print("\n  No reign parsing issues found!")

# ============================================================================
# SECTION 3: CATEGORY AUDIT
# ============================================================================
print("\n" + "=" * 100)
print("3. CATEGORY AUDIT")
print("=" * 100)

category_map = {}
for row in range(2, 120):
    cat = get_cell(row, "Category")
    if cat:
        normalized = normalize_category(cat)
        if cat not in category_map:
            category_map[cat] = normalized
        print(f"  '{cat}' -> '{normalized}'")

print(f"\n  Unique categories: {len(category_map)}")
for orig, norm in sorted(category_map.items()):
    print(f"    '{orig}' -> '{norm}'")

# ============================================================================
# SECTION 4: METAL/COMPOSITION AUDIT
# ============================================================================
print("\n" + "=" * 100)
print("4. METAL/COMPOSITION AUDIT")
print("=" * 100)

metal_map = {}
for row in range(2, 120):
    metal = get_cell(row, "Composition")
    if metal:
        normalized = normalize_metal(metal)
        if metal not in metal_map:
            metal_map[metal] = normalized
            
print(f"  Unique metals: {len(metal_map)}")
for orig, norm in sorted(metal_map.items()):
    status = "✓" if norm != "uncertain" else "⚠"
    print(f"    {status} '{orig}' -> '{norm}'")

# ============================================================================
# SECTION 5: REFERENCE PARSING AUDIT
# ============================================================================
print("\n" + "=" * 100)
print("5. REFERENCE PARSING AUDIT")
print("=" * 100)

ref_stats = {"total": 0, "parsed": 0, "other": 0}
other_refs = []

for row in range(2, 120):
    ref = get_cell(row, "Reference")
    if not ref:
        continue
    
    parsed = parse_references(ref)
    ref_stats["total"] += 1
    
    for p in parsed:
        if p["system"] == "other":
            ref_stats["other"] += 1
            other_refs.append((row, p["number"]))
        else:
            ref_stats["parsed"] += 1

print(f"  Total reference strings: {ref_stats['total']}")
print(f"  Successfully parsed references: {ref_stats['parsed']}")
print(f"  Fell back to 'other': {ref_stats['other']}")

if other_refs:
    print(f"\n  'Other' references (first 30):")
    for row, ref in other_refs[:30]:
        print(f"    Row {row}: '{ref}'")
    if len(other_refs) > 30:
        print(f"    ... and {len(other_refs) - 30} more")

# ============================================================================
# SECTION 6: OBVERSE/REVERSE LEGEND PARSING
# ============================================================================
print("\n" + "=" * 100)
print("6. OBVERSE/REVERSE PARSING AUDIT")
print("=" * 100)

legend_stats = {"with_legend": 0, "no_legend": 0, "total": 0}

for row in range(2, 120):
    obv = get_cell(row, "Obverse")
    if not obv:
        continue
    
    legend_stats["total"] += 1
    
    # Check if it has a legend (text before comma, all caps)
    if "," in obv:
        parts = obv.split(",", 1)
        potential_legend = parts[0].strip()
        # Check if it looks like a legend (mostly uppercase, under 60 chars)
        if len(potential_legend) < 60:
            legend_stats["with_legend"] += 1
        else:
            legend_stats["no_legend"] += 1
    else:
        legend_stats["no_legend"] += 1

print(f"  Total obverse descriptions: {legend_stats['total']}")
print(f"  With identifiable legend: {legend_stats['with_legend']}")
print(f"  Description only (no legend): {legend_stats['no_legend']}")

# ============================================================================
# SECTION 7: WEIGHT/DIAMETER AUDIT
# ============================================================================
print("\n" + "=" * 100)
print("7. WEIGHT/DIAMETER AUDIT")
print("=" * 100)

weight_issues = []
diameter_issues = []

for row in range(2, 120):
    weight = get_cell(row, "Weight")
    diameter = get_cell(row, "Diameter")
    ruler = get_cell(row, "Ruler Issuer")
    
    if weight:
        try:
            w = float(weight)
            if w <= 0 or w > 50:  # Suspicious weight
                weight_issues.append((row, ruler, weight, "OUT OF RANGE"))
        except:
            weight_issues.append((row, ruler, weight, "NOT A NUMBER"))
    
    if diameter:
        try:
            d = float(diameter)
            if d <= 0 or d > 50:  # Suspicious diameter
                diameter_issues.append((row, ruler, diameter, "OUT OF RANGE"))
        except:
            diameter_issues.append((row, ruler, diameter, "NOT A NUMBER"))

print(f"  Weight issues: {len(weight_issues)}")
for row, ruler, val, problem in weight_issues[:10]:
    print(f"    Row {row} ({ruler}): '{val}' - {problem}")

print(f"  Diameter issues: {len(diameter_issues)}")
for row, ruler, val, problem in diameter_issues[:10]:
    print(f"    Row {row} ({ruler}): '{val}' - {problem}")

# ============================================================================
# SECTION 8: NGC GRADE PARSING
# ============================================================================
print("\n" + "=" * 100)
print("8. NGC GRADE AUDIT")
print("=" * 100)

ngc_patterns = {}
for row in range(2, 120):
    ngc = get_cell(row, "NGC Grade")
    if ngc:
        if ngc not in ngc_patterns:
            ngc_patterns[ngc] = 0
        ngc_patterns[ngc] += 1

print(f"  Unique NGC grade patterns: {len(ngc_patterns)}")
for grade, count in sorted(ngc_patterns.items(), key=lambda x: -x[1]):
    print(f"    '{grade}' ({count}x)")

# ============================================================================
# SECTION 9: CHECK DB DATA INTEGRITY
# ============================================================================
print("\n" + "=" * 100)
print("9. DATABASE INTEGRITY CHECK")
print("=" * 100)

# Check for missing required fields
missing_fields = {"category": 0, "metal": 0, "denomination": 0, "issuing_authority": 0}
for coin in coins:
    if not coin.category:
        missing_fields["category"] += 1
    if not coin.metal:
        missing_fields["metal"] += 1
    if not coin.denomination:
        missing_fields["denomination"] += 1
    if not coin.issuing_authority:
        missing_fields["issuing_authority"] += 1

print(f"  Coins missing required fields:")
for field, count in missing_fields.items():
    status = "✓" if count == 0 else "⚠"
    print(f"    {status} {field}: {count}")

# Check references
ref_count = db.query(CoinReference).count()
ref_type_count = db.query(ReferenceType).count()
print(f"\n  Reference statistics:")
print(f"    CoinReference records: {ref_count}")
print(f"    ReferenceType records: {ref_type_count}")

# Check coins with/without references
coins_with_refs = db.query(Coin).join(CoinReference).distinct().count()
print(f"    Coins with references: {coins_with_refs}")
print(f"    Coins without references: {len(coins) - coins_with_refs}")

# ============================================================================
# SECTION 10: SAMPLE DATA COMPARISON
# ============================================================================
print("\n" + "=" * 100)
print("10. SAMPLE DATA COMPARISON (First 5 rows)")
print("=" * 100)

for row in range(2, 7):
    ruler = get_cell(row, "Ruler Issuer")
    denom = get_cell(row, "Coin type")
    minted = get_cell(row, "Minted")
    ref = get_cell(row, "Reference")
    
    # Find in DB
    db_coin = db.query(Coin).filter(
        Coin.issuing_authority == ruler,
        Coin.denomination == denom
    ).first()
    
    print(f"\n  Row {row}: {ruler} - {denom}")
    print(f"    Excel Minted: '{minted}'")
    if db_coin:
        print(f"    DB Years: {db_coin.mint_year_start} to {db_coin.mint_year_end}")
        print(f"    Excel Ref: '{ref[:50] if ref else 'N/A'}...'")
        
        # Get DB references
        db_refs = db.query(CoinReference).filter(CoinReference.coin_id == db_coin.id).all()
        for dbref in db_refs:
            ref_type = db.query(ReferenceType).filter(ReferenceType.id == dbref.reference_type_id).first()
            if ref_type:
                print(f"    DB Ref: {ref_type.system} {ref_type.local_ref}")
    else:
        print(f"    NOT FOUND IN DB!")

db.close()

# ============================================================================
# SUMMARY
# ============================================================================
print("\n" + "=" * 100)
print("AUDIT SUMMARY")
print("=" * 100)
print(f"  Total issues logged: {len(issues)}")
